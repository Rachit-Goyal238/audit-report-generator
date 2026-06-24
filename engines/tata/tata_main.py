import os
import json
import pandas as pd

from openpyxl import load_workbook

from engines.tata.report_utils import (
    create_output_paths
)

from engines.tata.pdf_utils import (
    extract_pdf_header,
    extract_evidence_pages,
    merge_pdfs,
    excel_to_pdf
)

from engines.tata.excel_utils import (
    populate_headers,
    populate_checklist
)

def generate_report(
    audit_id,
    master_file,
    client,
    template_type,
    pdf_file,
    annexure_pdf=None
):

    with open(
        "templates.json",
        "r",
        encoding="utf-8"
    ) as f:

        template_repository = json.load(f)

    template_file = template_repository[client][template_type]

    output_folder = "output"
    os.makedirs(output_folder, exist_ok=True)

    df = pd.read_excel(
        master_file,
        dtype=str,
        keep_default_na=False
    )

    df.columns = df.columns.str.strip()

    audit_df = df[
        df["Audit ID"].astype(str).str.strip() == audit_id
    ]

    if audit_df.empty:
        raise Exception(
            f"Audit ID '{audit_id}' not found"
        )

    print(f"Found {len(audit_df)} records")

    first_row = audit_df.iloc[0]

    agency_code = str(
        first_row["Agency Code"]
    ).strip()

    agency_name = str(
        first_row["Agency Name"]
    ).strip()

    paths = create_output_paths(
    agency_code,
    agency_name
    )

    generated_excel = paths["excel"]
    generated_pdf = paths["pdf"]
    evidence_pdf = paths["evidence"]
    final_report_pdf = paths["final"]

    pdf_data = extract_pdf_header(
        pdf_file
    )

    wb = load_workbook(
        template_file
    )

    checklist_sheet = wb.sheetnames[0]
    ws = wb[checklist_sheet]

    populate_headers(
    ws,
    first_row,
    pdf_data
)

    populate_checklist(
    ws,
    audit_df
)
    
    wb.save(
        generated_excel
    )

    print(generated_excel)
    
    print(
        "Excel workbook created"
    )

    print(generated_excel)
    print(os.path.exists(generated_excel))

    excel_to_pdf(
        generated_excel,
        generated_pdf
    )

    print(
        "PDF created"
    )

    extract_evidence_pages(
        pdf_file,
        evidence_pdf
    )

    print(
        "Evidence PDF created"
    )

    if annexure_pdf:

        merge_pdfs(
            generated_pdf,
            evidence_pdf,
            annexure_pdf,
            final_report_pdf
        )

    else:

        merge_pdfs(
            generated_pdf,
            evidence_pdf,
            None,
            final_report_pdf
        )

    print(
        "Final report created"
    )

    return {
        "excel": generated_excel,
        "pdf": generated_pdf,
        "evidence": evidence_pdf,
        "final": final_report_pdf
    }