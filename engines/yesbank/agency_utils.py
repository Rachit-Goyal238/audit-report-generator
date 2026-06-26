from openpyxl.styles import Alignment


def populate_agency(
    agency_ws,
    kaf_ws
):

    grouped = {}

    for row in range(
        2,
        kaf_ws.max_row + 1
    ):

        sr_no = kaf_ws.cell(
            row,
            1
        ).value

        observation = kaf_ws.cell(
            row,
            9
        ).value

        if (
            sr_no is None
            or
            observation in (
                None,
                ""
            )
        ):

            continue

        grouped.setdefault(
            sr_no,
            []
        ).append(
            str(observation).strip()
        )

    agency_lookup = {}

    for row in range(
        2,
        agency_ws.max_row + 1
    ):

        sr_no = agency_ws.cell(
            row,
            1
        ).value

        if sr_no is not None:

            agency_lookup[
                sr_no
            ] = row

    for sr_no, observations in grouped.items():

        if sr_no not in agency_lookup:

            continue

        row = agency_lookup[
            sr_no
        ]

        count = len(
            observations
        )

        if count > 4:

            errors = ">4"

        else:

            errors = count

        unique = []

        for obs in observations:

            if obs not in unique:

                unique.append(
                    obs
                )

        observation_text = "\n".join(
            unique
        )

        # Status Complied = No
        agency_ws.cell(
            row,
            8
        ).value = "No"

        agency_ws.cell(
            row,
            9
        ).value = errors

        obs_cell = agency_ws.cell(
            row,
            12
        )

        obs_cell.value = observation_text

        obs_cell.alignment = Alignment(
            wrap_text=True,
            vertical="top"
        )